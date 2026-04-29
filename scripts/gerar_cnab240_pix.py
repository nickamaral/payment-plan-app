#!/usr/bin/env python3
"""
Gerador de arquivo CNAB 240 PIX para Itaú SISPAG
Layout: v086, Forma de Pagamento 45 (PIX Transferência)
Manual: download.itau.com.br/bankline/sispag_cnab.pdf
"""

import sys
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Instale: pip install openpyxl --break-system-packages")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Helpers de formatação de campos
# ---------------------------------------------------------------------------

def num(value, size):
    """Campo numérico: zeros à esquerda, trunca à direita se maior."""
    s = str(value).strip().replace('.', '').replace(',', '').replace('-', '').replace('/', '').replace(' ', '')
    # Pega só dígitos
    s = ''.join(c for c in s if c.isdigit())
    return s[-size:].zfill(size) if len(s) >= size else s.zfill(size)

def alfa(value, size):
    """Campo alfanumérico: espaços à direita, trunca se maior. Converte para ASCII."""
    s = str(value or '').strip().upper()
    # Normaliza caracteres especiais (mantém ASCII)
    replacements = {
        'Á': 'A', 'À': 'A', 'Ã': 'A', 'Â': 'A',
        'É': 'E', 'Ê': 'E', 'Ó': 'O', 'Õ': 'O', 'Ô': 'O',
        'Í': 'I', 'Ú': 'U', 'Ç': 'C', 'Ñ': 'N',
    }
    for orig, rep in replacements.items():
        s = s.replace(orig, rep)
    return s[:size].ljust(size)

def valor_cnab(valor_reais, size=15):
    """Converte R$ para inteiro com 2 casas decimais implícitas."""
    v = float(str(valor_reais).replace(',', '.'))
    centavos = round(v * 100)
    return str(centavos).zfill(size)

def formatar_data(d):
    """Converte data (datetime, string) para DDMMAAAA."""
    if d is None or str(d).strip() in ('', 'None', 'nan'):
        return datetime.now().strftime("%d%m%Y")
    if hasattr(d, 'strftime'):
        return d.strftime("%d%m%Y")
    s = str(d).strip().replace('/', '').replace('-', '')
    if len(s) == 8:
        # Pode ser DDMMAAAA ou AAAAMMDD
        if int(s[4:]) > 1900:  # DDMMAAAA
            return s
        else:  # AAAAMMDD
            return s[6:8] + s[4:6] + s[0:4]
    return datetime.now().strftime("%d%m%Y")

def limpar_cnpj_cpf(value):
    """Remove pontuação de CPF/CNPJ."""
    return ''.join(c for c in str(value or '') if c.isdigit())


# ---------------------------------------------------------------------------
# Montagem dos registros CNAB 240
# ---------------------------------------------------------------------------

def header_arquivo(config, seq_arquivo=1):
    """Registro tipo 0 — Header de Arquivo (240 bytes)."""
    agora = datetime.now()
    cnpj = num(limpar_cnpj_cpf(config['cnpj']), 14)

    linha = (
        "341"                                +  # 001-003 Banco
        "0000"                               +  # 004-007 Lote
        "0"                                  +  # 008     Tipo Registro
        " " * 9                              +  # 009-017 Uso FEBRABAN
        "2"                                  +  # 018     Tipo Inscrição (CNPJ)
        cnpj                                 +  # 019-032 Número Inscrição
        " " * 20                             +  # 033-052 Código Convênio
        num(config['agencia'], 5)            +  # 053-057 Agência
        alfa(config.get('digito_agencia', ' '), 1) +  # 058 Dígito Agência
        num(config['conta'], 12)             +  # 059-070 Número Conta
        alfa(config.get('digito_conta', ' '), 1) +    # 071 Dígito Conta
        " "                                  +  # 072     Dígito Ag/Conta
        alfa(config['nome_empresa'], 30)     +  # 073-102 Nome Empresa
        alfa("BANCO ITAU SA", 30)            +  # 103-132 Nome Banco
        " " * 10                             +  # 133-142 Uso FEBRABAN
        "1"                                  +  # 143     Código Remessa
        agora.strftime("%d%m%Y")             +  # 144-151 Data Geração
        agora.strftime("%H%M%S")             +  # 152-157 Hora Geração
        str(seq_arquivo).zfill(6)            +  # 158-163 Sequencial Arquivo
        "080"                                +  # 164-166 Layout Arquivo (v080)
        "01600"                              +  # 167-171 Densidade
        " " * 20                             +  # 172-191 Uso Banco
        " " * 20                             +  # 192-211 Uso Empresa
        " " * 29                                # 212-240 Uso FEBRABAN
    )
    assert len(linha) == 240, f"ERRO Header Arquivo: {len(linha)} bytes"
    return linha


def header_lote(config, num_lote, data_credito_str, tipo_servico="30"):
    """Registro tipo 1 — Header de Lote (240 bytes)."""
    agora = datetime.now()
    cnpj = num(limpar_cnpj_cpf(config['cnpj']), 14)

    linha = (
        "341"                                +  # 001-003 Banco
        str(num_lote).zfill(4)              +  # 004-007 Número Lote
        "1"                                  +  # 008     Tipo Registro
        "C"                                  +  # 009     Tipo Operação (Crédito)
        tipo_servico                         +  # 010-011 Tipo Serviço (30=Salários)
        "45"                                 +  # 012-013 Forma Pagamento (PIX)
        "040"                                +  # 014-016 Layout Lote (v040)
        " "                                  +  # 017     Uso FEBRABAN
        "2"                                  +  # 018     Tipo Inscrição (CNPJ)
        cnpj                                 +  # 019-032 CNPJ Empresa
        " " * 20                             +  # 033-052 Código Convênio
        num(config['agencia'], 5)            +  # 053-057 Agência
        alfa(config.get('digito_agencia', ' '), 1) +  # 058 Dígito Agência
        num(config['conta'], 12)             +  # 059-070 Conta
        alfa(config.get('digito_conta', ' '), 1) +    # 071 Dígito Conta
        " "                                  +  # 072     Dígito Ag/Conta
        alfa(config['nome_empresa'], 30)     +  # 073-102 Nome Empresa
        " " * 40                             +  # 103-142 Informação 1
        " " * 30                             +  # 143-172 Informação 2
        "00001"                              +  # 173-177 Número Remessa
        agora.strftime("%d%m%Y")             +  # 178-185 Data Gravação
        data_credito_str                     +  # 186-193 Data Crédito
        " " * 47                                # 194-240 Uso FEBRABAN
    )
    assert len(linha) == 240, f"ERRO Header Lote: {len(linha)} bytes"
    return linha


def _agencia_conta_favorecido(p):
    """
    Monta 20 bytes de Agência/Conta conforme Nota 11a (Itaú) ou 11b (outros).
    Posições 024-043 do Segmento A.
    """
    banco = str(p.get('banco_favorecido', '341')).strip().zfill(3)

    if banco in ('341', '409'):
        # Nota 11a — layout proprietário Itaú
        # 024: zero | 025-028: agência 4d | 029: dígito ag | 030-035: zeros | 036-041: conta 6d | 042: dígito cc | 043: zero
        ag = num(str(p.get('agencia_favorecido', '0000')), 4)
        dig_ag = alfa(p.get('digito_agencia_fav', '0'), 1)
        cc = num(str(p.get('conta_favorecido', '000000')), 6)
        dig_cc = alfa(p.get('digito_conta_fav', '0'), 1)
        return "0" + ag + dig_ag + "000000" + cc + dig_cc + "0"
    else:
        # Nota 11b — padrão FEBRABAN
        ag = num(str(p.get('agencia_favorecido', '00000')), 5)
        dig_ag = alfa(p.get('digito_agencia_fav', ' '), 1)
        cc = num(str(p.get('conta_favorecido', '000000000000')), 12)
        dig_cc = alfa(p.get('digito_conta_fav', ' '), 1)
        return ag + dig_ag + cc + dig_cc + " "


def segmento_a(num_lote, seq, p):
    """
    Registro tipo 3, Segmento A — dados de crédito PIX (240 bytes).
    Campo 113-114 (dentro de Qtde Moeda 105-119) carrega o Tipo de Transferência PIX.
    """
    tipo_pix = str(p.get('tipo_pix', '01')).strip().upper()
    banco_fav = str(p.get('banco_favorecido', '341')).strip().zfill(3)

    # Se pagamento por chave PIX (tipo 04), zera banco/agência/conta no Seg A
    if tipo_pix == '04':
        banco_str = "000"
        ag_cc = "0" * 20
    else:
        banco_str = banco_fav
        ag_cc = _agencia_conta_favorecido(p)

    # Quantidade Moeda (pos 105-119, 15 bytes):
    # 8 zeros | tipo PIX (2 bytes) | 5 zeros
    qtde_moeda = "00000000" + tipo_pix.ljust(2) + "00000"

    num_doc = alfa(str(p.get('num_documento', '')), 20)
    data_pag = formatar_data(p.get('data_pagamento'))

    linha = (
        "341"                              +  # 001-003 Banco
        str(num_lote).zfill(4)            +  # 004-007 Lote
        "3"                                +  # 008     Tipo Registro
        str(seq).zfill(5)                 +  # 009-013 Sequência no Lote
        "A"                                +  # 014     Segmento
        "0"                                +  # 015     Tipo Movimento (0=Inclusão)
        "00"                               +  # 016-017 Código Instrução (Sem Instrução)
        "009"                              +  # 018-020 Câmara (SPI)
        banco_str                          +  # 021-023 Banco Favorecido
        ag_cc                              +  # 024-043 Agência/Conta (20 bytes)
        alfa(p['nome_favorecido'], 30)     +  # 044-073 Nome Favorecido
        num_doc                            +  # 074-093 Nº Documento Empresa
        data_pag                           +  # 094-101 Data Pagamento (DDMMAAAA)
        "BRL"                              +  # 102-104 Tipo Moeda
        qtde_moeda                         +  # 105-119 Qtde Moeda (c/ tipo PIX em 113-114)
        valor_cnab(p['valor'], 15)         +  # 120-134 Valor Pagamento
        " " * 20                           +  # 135-154 Nº Doc Banco
        "0" * 8                            +  # 155-162 Data Real Efetivação
        "0" * 15                           +  # 163-177 Valor Real Efetivação
        " " * 40                           +  # 178-217 Informação 2 (uso livre)
        "0"                                +  # 218     Aviso Favorecido
        " " * 12                           +  # 219-230 Ocorrências (retorno)
        " " * 10                              # 231-240 Uso FEBRABAN
    )
    assert len(linha) == 240, f"ERRO Segmento A seq {seq}: {len(linha)} bytes"
    return linha


def segmento_b(num_lote, seq, p):
    """
    Registro tipo 3, Segmento B — chave PIX e dados complementares (240 bytes).
    Obrigatório quando tipo_pix = '04'. Opcional (mas recomendado) nos demais casos.
    """
    tipo_chave = str(p.get('tipo_chave', '03')).strip().zfill(2)
    cpf_cnpj = num(limpar_cnpj_cpf(p.get('cpf_cnpj_favorecido', '')), 14)
    txid = alfa(str(p.get('txid', '')), 30)        # v086: pos 033-062
    mensagem = alfa(str(p.get('mensagem', '')), 65)  # pos 063-127
    chave_pix = str(p.get('chave_pix', '') or '').strip()[:100].ljust(100)  # pos 128-227

    linha = (
        "341"              +  # 001-003 Banco
        str(num_lote).zfill(4) +  # 004-007 Lote
        "3"                +  # 008     Tipo Registro
        str(seq).zfill(5) +  # 009-013 Sequência no Lote
        "B"                +  # 014     Segmento
        tipo_chave         +  # 015-016 Tipo de Chave PIX
        "  "               +  # 017-018 Uso FEBRABAN
        cpf_cnpj           +  # 019-032 CPF/CNPJ Favorecido
        txid               +  # 033-062 TXID (v086)
        mensagem           +  # 063-127 Mensagem (Informação entre Usuários)
        chave_pix          +  # 128-227 Chave PIX (100 bytes)
        " " * 13              # 228-240 Uso FEBRABAN
    )
    assert len(linha) == 240, f"ERRO Segmento B seq {seq}: {len(linha)} bytes"
    return linha


def trailer_lote(num_lote, qtd_registros, valor_total):
    """Registro tipo 5 — Trailer de Lote (240 bytes)."""
    linha = (
        "341"                               +  # 001-003 Banco
        str(num_lote).zfill(4)             +  # 004-007 Lote
        "5"                                 +  # 008     Tipo Registro
        " " * 9                             +  # 009-017 Uso FEBRABAN
        str(qtd_registros).zfill(6)        +  # 018-023 Qtde Registros no Lote
        valor_cnab(valor_total, 18)         +  # 024-041 Valor Total do Lote
        "0" * 6                             +  # 042-047 Qtde Moeda
        " " * 193                              # 048-240 Uso FEBRABAN
    )
    assert len(linha) == 240, f"ERRO Trailer Lote: {len(linha)} bytes"
    return linha


def trailer_arquivo(qtd_lotes, qtd_registros_total):
    """Registro tipo 9 — Trailer de Arquivo (240 bytes)."""
    linha = (
        "341"                               +  # 001-003 Banco
        "9999"                              +  # 004-007 Lote
        "9"                                 +  # 008     Tipo Registro
        " " * 9                             +  # 009-017 Uso FEBRABAN
        str(qtd_lotes).zfill(6)            +  # 018-023 Qtde Lotes
        str(qtd_registros_total).zfill(6)  +  # 024-029 Qtde Registros no Arquivo
        "0" * 6                             +  # 030-035 Qtde Contas Conciliação
        " " * 205                              # 036-240 Uso FEBRABAN
    )
    assert len(linha) == 240, f"ERRO Trailer Arquivo: {len(linha)} bytes"
    return linha


# ---------------------------------------------------------------------------
# Leitura da planilha Excel
# ---------------------------------------------------------------------------

def _normalizar(s):
    """Remove acentos e caracteres especiais para comparação de chaves."""
    replacements = {
        'Á': 'A', 'À': 'A', 'Ã': 'A', 'Â': 'A',
        'É': 'E', 'Ê': 'E', 'Ó': 'O', 'Õ': 'O', 'Ô': 'O',
        'Í': 'I', 'Ú': 'U', 'Ç': 'C', 'Ñ': 'N',
    }
    s = s.upper()
    for orig, rep in replacements.items():
        s = s.replace(orig, rep)
    return s

def ler_config(ws_config):
    """Lê aba Configuração (rótulo na coluna B, valor na coluna C)."""
    config = {}
    chave_map = {
        'CNPJ DA EMPRESA':        'cnpj',
        'NOME DA EMPRESA':        'nome_empresa',
        'AGENCIA':                'agencia',
        'DIGITO DA AGENCIA':      'digito_agencia',
        'CONTA':                  'conta',
        'DIGITO DA CONTA':        'digito_conta',
        'TIPO DE SERVICO':        'tipo_servico',
        'DATA DE PAGAMENTO':      'data_pagamento',
    }
    for row in ws_config.iter_rows(min_row=4, values_only=True):
        if row[1] and row[2] is not None:
            chave_raw = _normalizar(str(row[1]).strip())
            for k, v in chave_map.items():
                if k in chave_raw:
                    config[v] = row[2]
    return config


def ler_pagamentos(ws_pag):
    """Lê aba Pagamentos a partir da linha 4 (linha 3 = cabeçalho)."""
    pagamentos = []
    for row in ws_pag.iter_rows(min_row=4, values_only=True):
        if not row[0] or str(row[0]).strip() in ('', 'None', 'nan'):
            continue
        # Ignora linhas sem valor (ex: rodapé/legendas)
        if row[10] is None or str(row[10]).strip() in ('', 'None', 'nan'):
            continue
        pagamentos.append({
            'nome_favorecido':       row[0],
            'cpf_cnpj_favorecido':   row[1],
            'banco_favorecido':      row[2],
            'agencia_favorecido':    row[3],
            'digito_agencia_fav':    row[4],
            'conta_favorecido':      row[5],
            'digito_conta_fav':      row[6],
            'tipo_pix':              row[7],   # 01/PG/03/04
            'tipo_chave':            row[8],   # 01/02/03/04 (só se tipo_pix=04)
            'chave_pix':             row[9],
            'valor':                 row[10],
            'data_pagamento':        row[11],
            'num_documento':         row[12],
            'mensagem':              row[13],
        })
    return pagamentos


# ---------------------------------------------------------------------------
# Geração do arquivo CNAB 240
# ---------------------------------------------------------------------------

def gerar_cnab240(planilha_path, arquivo_saida=None):
    wb = openpyxl.load_workbook(planilha_path, data_only=True)

    config = ler_config(wb['Configuração'])
    pagamentos = ler_pagamentos(wb['Pagamentos'])

    if not pagamentos:
        print("Nenhum pagamento encontrado na planilha.")
        sys.exit(1)

    # Valida campos obrigatórios na config
    for campo in ('cnpj', 'nome_empresa', 'agencia', 'conta'):
        if campo not in config:
            print(f"ERRO: Campo '{campo}' não encontrado na aba Configuração.")
            sys.exit(1)

    # Data de crédito: usa a da primeira linha de pagamento ou da config
    data_cred = config.get('data_pagamento') or pagamentos[0].get('data_pagamento')
    data_cred_str = formatar_data(data_cred)

    tipo_servico = str(config.get('tipo_servico', '30')).strip().zfill(2)
    num_lote = 1

    linhas = []

    # --- Header Arquivo ---
    linhas.append(header_arquivo(config))

    # --- Header Lote ---
    linhas.append(header_lote(config, num_lote, data_cred_str, tipo_servico))

    seq = 1
    valor_total = 0.0
    qtd_detalhes = 0

    for pag in pagamentos:
        # Segmento A (obrigatório)
        linhas.append(segmento_a(num_lote, seq, pag))
        seq += 1
        qtd_detalhes += 1

        # Segmento B: obrigatório se tipo_pix=04, recomendado nos demais
        tipo_pix = str(pag.get('tipo_pix', '01')).strip().upper()
        tem_chave = str(pag.get('chave_pix', '') or '').strip()
        if tipo_pix == '04' or tem_chave:
            linhas.append(segmento_b(num_lote, seq, pag))
            seq += 1
            qtd_detalhes += 1

        valor_total += float(str(pag['valor']).replace(',', '.'))

    # Qtde de registros no lote = header_lote + detalhes + trailer_lote
    qtd_reg_lote = 1 + qtd_detalhes + 1

    # --- Trailer Lote ---
    linhas.append(trailer_lote(num_lote, qtd_reg_lote, valor_total))

    # Qtde total no arquivo = header_arquivo + tudo no lote + trailer_arquivo
    qtd_total = 1 + qtd_reg_lote + 1

    # --- Trailer Arquivo ---
    linhas.append(trailer_arquivo(1, qtd_total))

    # Define nome do arquivo de saída
    if not arquivo_saida:
        base = os.path.splitext(planilha_path)[0]
        arquivo_saida = base + "_CNAB240.txt"

    # Modo binário garante CRLF exato sem dupla-tradução do Python
    with open(arquivo_saida, 'wb') as f:
        for linha in linhas:
            f.write((linha + '\r\n').encode('ascii', errors='replace'))

    print(f"\n✓ Arquivo gerado: {arquivo_saida}")
    print(f"  Registros totais : {qtd_total}")
    print(f"  Pagamentos       : {len(pagamentos)}")
    print(f"  Valor total      : R$ {valor_total:,.2f}")
    print(f"\n  Próximo passo: valide em")
    print(f"  Itaú Empresas > Transmissão de Arquivos > Validação > Layout de Arquivo")

    return arquivo_saida


# ---------------------------------------------------------------------------
# Verificador de integridade
# ---------------------------------------------------------------------------

def verificar_arquivo(path):
    print(f"\nVerificando: {path}")
    erros = []
    with open(path, 'rb') as f:
        raw = f.read()
    linhas_raw = raw.split(b'\r\n')
    for i, linha in enumerate(linhas_raw, 1):
        if not linha:
            continue  # ignora última linha vazia após split final
        if len(linha) != 240:
            tipo = chr(linha[7]) if len(linha) > 7 else '?'
            erros.append(f"  Linha {i}: {len(linha)} bytes (esperado 240) — tipo='{tipo}'")

    if erros:
        print(f"PROBLEMAS ENCONTRADOS ({len(erros)}):")
        for e in erros:
            print(e)
    else:
        print(f"✓ Todas as linhas têm exatamente 240 bytes.")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python gerar_cnab240_pix.py <planilha.xlsx> [saida.txt]")
        sys.exit(1)

    planilha = sys.argv[1]
    saida = sys.argv[2] if len(sys.argv) > 2 else None

    arquivo = gerar_cnab240(planilha, saida)
    verificar_arquivo(arquivo)
